import smtplib
import openpyxl

def main():
    print("Welcome! With this application you will be able to send emails")
    send_email()
    # manage_excel()
    
def manage_excel():
    book = openpyxl.load_workbook('tracker.xlsx')
    sheet = book['Hoja1']
    rows = sheet.max_row
    columns = sheet.max_column
    print(f"Rows: {rows}\nColumns: {columns}")
    
    print("ID|Email|Class 1|Class 2|Class 3|")
    for row in (1,rows):
        for column in (1,columns):
            print(sheet.cell(row,column).value)
            
    
def send_email():
    smtp_server = 'smtp.gmail.com'
    smtp_port = 587
    smtp_username = 'danieljuegos1125@gmail.com'
    smtp_password = 'cuipzdbptzcrckev'
    from_email = smtp_username
    to_email = 'gabriela14mejia@gmail.com'
    subject = 'Te amo!'
    body = 'Este correo es para el amor de mi vida. Espero que tengas un dia muy lindo y lleno de amor. Eres mi mayor inspiracion!'

    message = f'Subject: {subject}\n\n{body}'

    with smtplib.SMTP(smtp_server,smtp_port) as smtp:
        smtp.starttls()
        smtp.login(smtp_username,smtp_password)
        smtp.sendmail(from_email,to_email,message)

if __name__ == '__main__':
    main()